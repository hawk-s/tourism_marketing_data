"""
to_excel.py
Transforms booking_hotels_data.json into a structured Excel workbook.

Sheets produced:
  1. Hotels          – one row per hotel, flat fields + subscores as columns
  2. Facilities      – one row per facility item (from facilities_grouped)
  3. Area_Info       – one row per nearby place / transport entry
"""

import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths relative to this script file so it works from any working directory
_HERE       = Path(__file__).parent
INPUT_FILE  = _HERE / "booking_hotels_data.json"
OUTPUT_FILE = _HERE / "booking_hotels_data.xlsx"

# ── Load data ──────────────────────────────────────────────────────────────────
with INPUT_FILE.open(encoding="utf-8") as f:
    hotels = json.load(f)

print(f"Loaded {len(hotels)} hotels from {INPUT_FILE}")

# ── Helpers ────────────────────────────────────────────────────────────────────

def pipe(lst):
    """Join a list into a pipe-separated string, or return None."""
    if not lst:
        return None
    cleaned = [str(x).strip() for x in lst if x is not None and str(x).strip()]
    return " | ".join(cleaned) if cleaned else None


def dedupe_ordered(lst):
    """Remove duplicates while preserving order."""
    seen = set()
    out  = []
    for x in lst:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


# ── Sheet 1: Hotels ────────────────────────────────────────────────────────────
# Collect all subscore keys across all hotels so we can make consistent columns
all_subscore_keys = []
seen_keys = set()
for h in hotels:
    for k in (h.get("subscores") or {}).keys():
        if k not in seen_keys:
            all_subscore_keys.append(k)
            seen_keys.add(k)

hotels_rows = []
for h in hotels:
    subscores = h.get("subscores") or {}
    row = {
        "hotel_id":             h.get("hotel_id"),
        "name":                 h.get("name"),
        "city":                 h.get("city"),
        "region":               h.get("region"),
        "country":              h.get("country"),
        "dest_name":            h.get("dest_name"),
        "star_rating":          h.get("star_rating"),
        "property_type":        h.get("property_type"),
        "overall_score":        h.get("overall_score"),
        "score_label":          h.get("score_label"),
        "num_reviews":          h.get("num_reviews"),
    }
    # Subscores as individual columns
    for k in all_subscore_keys:
        row[f"score_{k.lower().replace(' ', '_')}"] = subscores.get(k)

    row.update({
        "description":              h.get("description"),
        "highlights":               pipe(h.get("highlights")),
        "most_popular_facilities":  pipe(dedupe_ordered(h.get("most_popular_facilities") or [])),
        "facilities":               pipe(h.get("facilities")),
        "sustainability":           h.get("sustainability"),
        "awards":                   pipe(h.get("awards")),
        "lat":                      h.get("lat"),
        "lng":                      h.get("lng"),
        "scraped_at":               h.get("scraped_at"),
        "url":                      h.get("url"),
    })
    hotels_rows.append(row)

df_hotels = pd.DataFrame(hotels_rows)

# ── Sheet 2: Facilities (normalised from facilities_grouped) ───────────────────
facility_rows = []
for h in hotels:
    hotel_id   = h.get("hotel_id")
    hotel_name = h.get("name")
    fg = h.get("facilities_grouped") or {}
    for category, data in fg.items():
        note  = (data or {}).get("note")
        items = (data or {}).get("items") or []
        for item in items:
            facility_rows.append({
                "hotel_id":          hotel_id,
                "hotel_name":        hotel_name,
                "category":          category,
                "category_note":     note,
                "facility":          item.get("name"),
                "additional_charge": item.get("additional_charge"),
            })

df_facilities = pd.DataFrame(facility_rows)

# ── Sheet 3: Area Info (normalised) ───────────────────────────────────────────
area_rows = []
for h in hotels:
    hotel_id   = h.get("hotel_id")
    hotel_name = h.get("name")
    ai = h.get("area_info") or {}
    for section, items in ai.items():
        for item in (items or []):
            area_rows.append({
                "hotel_id":   hotel_id,
                "hotel_name": hotel_name,
                "section":    section,
                "name":       item.get("name"),
                "type":       item.get("type"),
                "distance_m": item.get("distance_m"),
            })

df_area = pd.DataFrame(area_rows)

# ── Write to Excel ─────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_hotels.to_excel(writer,     sheet_name="Hotels",     index=False)
    df_facilities.to_excel(writer, sheet_name="Facilities", index=False)
    df_area.to_excel(writer,       sheet_name="Area_Info",  index=False)

# ── Style ──────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN    = Alignment(vertical="top", wrap_text=False)
WRAP_ALIGN    = Alignment(vertical="top", wrap_text=True)
THIN          = Side(style="thin", color="D0D0D0")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns in the Hotels sheet that contain long text and should wrap + be wider
WRAP_COLS = {"description", "highlights", "most_popular_facilities", "facilities"}

wb = load_workbook(OUTPUT_FILE)

for sheet_name in ["Hotels", "Facilities", "Area_Info"]:
    ws = wb[sheet_name]

    # Freeze top row
    ws.freeze_panes = "A2"

    # Style header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = BORDER

    # Auto-fit column widths + style data cells
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        col_letter  = get_column_letter(col_idx)
        header_val  = str(col_cells[0].value or "")
        is_wrap_col = header_val in WRAP_COLS

        max_len = len(header_val)
        for cell in col_cells[1:]:
            val = cell.value
            cell.border    = BORDER
            cell.alignment = WRAP_ALIGN if is_wrap_col else CELL_ALIGN
            if val is not None:
                # For width calc, don't let long text cells dominate
                line_len = min(len(str(val).split("\n")[0]), 60)
                max_len  = max(max_len, line_len)

        # Cap widths
        if is_wrap_col:
            ws.column_dimensions[col_letter].width = 50
        else:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE}")
print(f"  Hotels sheet:     {len(df_hotels)} rows × {len(df_hotels.columns)} columns")
print(f"  Facilities sheet: {len(df_facilities)} rows × {len(df_facilities.columns)} columns")
print(f"  Area_Info sheet:  {len(df_area)} rows × {len(df_area.columns)} columns")
